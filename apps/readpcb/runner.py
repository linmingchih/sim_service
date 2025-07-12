"""Convert a BRD layout to AEDB and export its stackup table."""
import argparse
import os
import zipfile
import openpyxl
from openpyxl import Workbook
from pyedb import Edb


def export_stackup(edb_obj, xlsx_path):
    data = []
    for layer_name, layer in edb_obj.stackup.stackup_layers.items():
        m = edb_obj.materials.materials[layer.material]
        if layer.type == 'signal':
            permittivity = ''
            loss_tangent = ''
            conductivity = m.conductivity
        else:
            permittivity = m.permittivity
            loss_tangent = m.dielectric_loss_tangent
            conductivity = ''
        thickness_mm = layer.thickness * 1000.0
        data.append([
            layer_name,
            layer.type,
            thickness_mm,
            permittivity,
            loss_tangent,
            conductivity
        ])
    wb = Workbook()
    ws = wb.active
    ws.title = "Stackup"
    header = ['Layer Name', 'Type', 'Thickness (mm)', 'Permittivity', 'Loss Tangent', 'Conductivity (S/m)']
    ws.append(header)
    for row in data:
        ws.append(row)
    wb.save(xlsx_path)


def table_html(xlsx_path, html_path):
    """Generate a styled HTML table from the stackup Excel sheet."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Stackup"]

    rows = []
    for idx, r in enumerate(ws.iter_rows(values_only=True)):
        tag = "th" if idx == 0 else "td"
        cells = "".join(f"<{tag}>{c}</{tag}>" for c in r)
        rows.append(f"<tr>{cells}</tr>")

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset='UTF-8'>
<style>
html, body {{
  height: 100%;
  margin: 0;
}}
body {{
  display: flex;
  justify-content: center;
  align-items: center;
  background: #f8f9fa;
}}
table {{
  border-collapse: collapse;
  box-shadow: 0 2px 6px rgba(0,0,0,0.1);
}}
th, td {{
  border: 1px solid #ccc;
  padding: 8px 12px;
  text-align: center;
}}
tr:nth-child(even) {{
  background-color: #d0f0ff;
}}
tr:nth-child(odd) {{
  background-color: #d0ffd6;
}}
</style>
</head>
<body>
<table>
{''.join(rows)}
</table>
</body>
</html>"""

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)


def main(brd_file):
    edb_name = 'board.aedb'
    edb = Edb(brd_file, edbversion='2024.1')
    export_stackup(edb, 'stackup.xlsx')
    # Save the project inside the current working directory so it can be
    # included in the output files instead of writing next to the source ``.brd``.
    edb.save_edb_as(edb_name)
    edb.close_edb()
    zip_name = 'board_aedb.zip'
    with zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(edb_name):
            for f in files:
                p = os.path.join(root, f)
                arc = os.path.relpath(p, edb_name)
                z.write(p, os.path.join('board.aedb', arc))
    table_html('stackup.xlsx', 'result.html')
    print(f"Generated {zip_name} and stackup.xlsx")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Convert BRD to AEDB and export stackup.')
    parser.add_argument('--brd', required=True, help='Input BRD file')
    args = parser.parse_args()
    main(args.brd)
