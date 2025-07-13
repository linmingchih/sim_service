"""Apply stackup changes from an Excel file to an AEDB archive."""
import argparse
import os
import shutil
import zipfile
import tempfile
import openpyxl
from openpyxl import Workbook
from pyedb import Edb


def export_stackup(edb_obj, xlsx_path):
    data = []
    for layer_name, layer in edb_obj.stackup.stackup_layers.items():
        m = edb_obj.materials.materials[layer.material]
        if layer.type == 'signal':
            permittivity = ''
            loss_tangent = ''
            conductivity = m.conductivity
        else:
            permittivity = m.permittivity
            loss_tangent = m.dielectric_loss_tangent
            conductivity = ''
        thickness_mm = layer.thickness * 1000.0
        data.append([
            layer_name,
            layer.type,
            thickness_mm,
            permittivity,
            loss_tangent,
            conductivity
        ])
    wb = Workbook()
    ws = wb.active
    ws.title = "Stackup"
    header = ['Layer Name', 'Type', 'Thickness (mm)', 'Permittivity', 'Loss Tangent', 'Conductivity (S/m)']
    ws.append(header)
    for row in data:
        ws.append(row)
    wb.save(xlsx_path)



def apply_xlsx(xlsx_path, edb_path, version):
    edb = Edb(edb_path, edbversion=version)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Stackup"]
    
    material_dic = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        layer_name, layer_type, thickness_mm, permittivity, loss_tangent, conductivity = row
        thickness_m = float(thickness_mm) / 1000.0
        edb.stackup.stackup_layers[layer_name].thickness = thickness_m
    
        if layer_type == "signal":
            if conductivity not in material_dic:
                name = f'metal_{conductivity}'
                mat = edb.materials.add_conductor_material(name, conductivity)
                material_dic[conductivity] = mat
                
            edb.stackup.stackup_layers[layer_name].material = material_dic[conductivity].name
        else:
            if (permittivity, loss_tangent) not in material_dic:
                name = f'dielectric_{permittivity}_{loss_tangent}'
                mat = edb.materials.add_dielectric_material(name, 
                                                            permittivity, 
                                                            loss_tangent)
                material_dic[(permittivity, loss_tangent)] = mat
            
            edb.stackup.stackup_layers[layer_name].material = material_dic[(permittivity, loss_tangent)].name
    
    edb.save_edb_as('modified.aedb')
    edb.close_edb()




def table_html(xlsx_path, html_path):
    """Generate a styled HTML table from the stackup Excel sheet."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Stackup"]

    rows = []
    for idx, r in enumerate(ws.iter_rows(values_only=True)):
        tag = "th" if idx == 0 else "td"
        cells = "".join(f"<{tag}>{c}</{tag}>" for c in r)
        rows.append(f"<tr>{cells}</tr>")

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset='UTF-8'>
<style>
html, body {{
  height: 100%;
  margin: 0;
}}
body {{
  display: flex;
  justify-content: center;
  align-items: center;
  background: #f8f9fa;
}}
table {{
  border-collapse: collapse;
  box-shadow: 0 2px 6px rgba(0,0,0,0.1);
}}
th, td {{
  border: 1px solid #ccc;
  padding: 8px 12px;
  text-align: center;
}}
tr:nth-child(even) {{
  background-color: #d0f0ff;
}}
tr:nth-child(odd) {{
  background-color: #d0ffd6;
}}
</style>
</head>
<body>
<table>
{''.join(rows)}
</table>
</body>
</html>"""

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)


def main(aedb_zip, xlsx_file):
    with tempfile.TemporaryDirectory() as tmp:
        with zipfile.ZipFile(aedb_zip) as z:
            z.extractall(tmp)
        aedb_dirs = [p for p in os.listdir(tmp) if p.endswith('.aedb')]
        if not aedb_dirs:
            raise FileNotFoundError(
                'No .aedb folder found in the provided zip archive.'
            )
        aedb_dir = aedb_dirs[0]
        aedb_path = os.path.join(tmp, aedb_dir)
        shutil.copy(xlsx_file, os.path.join(tmp, 'stackup.xlsx'))
        apply_xlsx(os.path.join(tmp, 'stackup.xlsx'), aedb_path)
        export_stackup(Edb(aedb_path, edbversion='2024.1'), os.path.join(tmp, 'updated.xlsx'))
        output_zip = 'updated_aedb.zip'
        with zipfile.ZipFile(output_zip, 'w', zipfile.ZIP_DEFLATED) as out:
            for root, _, files in os.walk(aedb_path):
                for f in files:
                    p = os.path.join(root, f)
                    arc = os.path.relpath(p, aedb_path)
                    out.write(p, os.path.join(aedb_dir, arc))
        shutil.copy(os.path.join(tmp, 'updated.xlsx'), 'updated.xlsx')
    table_html('updated.xlsx', 'result.html')
    print(f"Updated AEDB written to {output_zip}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Update stackup from Excel file.')
    parser.add_argument('--aedb_zip', required=True, help='Zipped AEDB')
    parser.add_argument('--xlsx', required=True, help='Stackup Excel file')
    args = parser.parse_args()
    main(args.aedb_zip, args.xlsx)
